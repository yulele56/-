# -*- coding: utf-8 -*-
"""
从流水文件夹读取各银行账户明细，区分收入/支出，按收款渠道分组输出。

支持格式:
    - 建行/农行/兴业/光大/华兴/邮政等 Excel 双列收支
    - 中行双语（交易类型 + 交易金额）
    - 中行 CSV 导出（utf-16 tab 分隔，交易日期 + 收入/支出金额）
    - 支付宝账务明细 CSV（# 开头说明行，skiprows=4，支出列为负数）
    - 微信支付账单 xlsx（交易时间 / 收/支 / 金额(元)）
    - 华兴网银 [HISTORYDETAIL] 导出（转入金额=收入、转出金额=支出；文件名可能含「工行」）
    - 广州银行等（借方金额 / 贷方金额）
    - 民生等 HTML 伪装 xls（read_html 兜底）

RPA 调用:
    df_all, result = get_bank_flows_from_folder(流水文件夹路径)
    result 顺序: 先各渠道支出明细，再各渠道收入明细

    # 批量写入对账台账（替代 RPA 逐行填 Excel）:
    written, total = append_flows_to_ledger(台账路径, 流水文件夹路径)
"""

import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

FLOW_SHEET_NAMES = (
    "收入流水汇总",
    "银行收入流水汇总",
    "银行收入流水",
    "收入流水",
)


def get_bank_flows_from_folder(flow_folder):
    """
    读取流水文件夹下各银行流水文件，解析为台账「银行收入流水汇总」所需字段并分组。

    参数:
        flow_folder: 流水根目录，递归扫描 *.xls / *.xlsx / *.csv

    字段说明:
        日期: 记账日期或交易日期，格式 YYYY-MM-DD
        对方账户: 对方名称
        收入: 收入金额（支出行留空字符串，不写 NaN）
        支出: 支出金额（收入行留空字符串）
        收款渠道: 由文件名识别的银行/账户渠道名
        收支类型: "收入" 或 "支出"
        备注: 摘要含往来款时写「往来款」，否则空串

    返回:
        df_all: 全部流水 DataFrame（供核对或写入台账）
        result: 分组列表，供 RPA 按序处理（按收款渠道 + 收支类型 + 日期）
            ① 支出组: {"收款渠道", "收支类型": "支出", "日期", "明细"}
            ② 收入组: {"收款渠道", "收支类型": "收入", "日期", "明细"}
    """
    # 各银行表头别名，用于自动匹配列名（子串匹配）
    date_keys = ("记账日期", "交易日期", "交易时间", "发生时间", "日期")
    income_keys = (
        "收入金额（+元）", "收入金额(+元)", "贷方发生额（收入）",
        "贷方发生额/元(收入)", "贷方金额(收入)", "贷方发生额", "贷方金额",
        "存入金额", "收入金额", "收入", "转入金额", "转入额",
    )
    expense_keys = (
        "支出金额（-元）", "支出金额(-元)", "借方发生额（支取）",
        "借方发生额/元(支取)", "借方金额(支出)", "借方发生额", "借方金额",
        "支出金额", "支出", "转出金额", "转出额",
    )
    counterparty_keys = (
        "对方户名", "对方名称", "对方账户", "对方账户名称",
        "对方单位名称", "对方单位", "收款人名称", "付款人名称",
    )
    zhaiyao_keys = (
        "业务摘要", "摘要", "业务类型", "客户附言", "附言", "用途",
    )

    def normalize_header(value):
        """表头单元格：去空白、换行，便于与 date_keys 等匹配。"""
        if pd.isna(value):
            return ""
        return str(value).strip().replace("\n", "")

    def clean_text(value):
        """对方账户等文本字段，空值转空串；重复列名时取首列。"""
        if isinstance(value, pd.Series):
            value = value.iloc[0] if not value.empty else ""
        if pd.isna(value):
            return ""
        return str(value).strip()

    def row_cell(row, col, default=""):
        """从行中取单元格；列名重复时取第一列标量。"""
        if not col:
            return default
        val = row.get(col, default)
        if isinstance(val, pd.Series):
            val = val.iloc[0] if not val.empty else default
        return val

    def row_summary_text(row, df):
        """合并行内摘要类列文本；华兴/工行等同时有「摘要」「附言」时均参与匹配（摘要优先）。"""
        texts = []
        used_cols = set()
        for key in zhaiyao_keys:
            for col in df.columns:
                if col in used_cols:
                    continue
                if key in str(col):
                    used_cols.add(col)
                    text = clean_text(row_cell(row, col))
                    if text:
                        texts.append(text)
                    break
        return " ".join(texts)

    def normalize_amount(value, expense_negative=False):
        """
        金额清洗：去逗号/货币符号，保留两位小数；无效或 ≤0 返回 None。
        expense_negative: 支付宝等导出在「支出金额（-元）」列写负数，取绝对值作为支出。
        """
        if isinstance(value, pd.Series):
            value = value.iloc[0] if not value.empty else None
        if pd.isna(value):
            return None
        text = str(value).strip().replace("\t", "")
        if text in ("", "-", "__", "nan", "NaN", "None"):
            return None
        text = text.replace(",", "").replace(" ", "").replace("￥", "").replace("¥", "").replace("元", "")
        try:
            amount = round(float(text), 2)
        except ValueError:
            return None
        if expense_negative and amount < 0:
            amount = abs(amount)
        return amount if amount > 0 else None

    def normalize_date(value):
        """日期统一为 YYYY-MM-DD；支持 20260420、20260420.0、带时分秒等。"""
        if pd.isna(value):
            return None
        text = str(value).strip()
        if text in ("", "-", "__", "nan"):
            return None
        if re.fullmatch(r"\d{8}\.0+", text):
            text = text.split(".")[0]
        if re.fullmatch(r"\d{8}", text):
            return "%s-%s-%s" % (text[0:4], text[4:6], text[6:8])
        if re.fullmatch(r"\d{8}\s+\d{6}", text):
            return "%s-%s-%s" % (text[0:4], text[4:6], text[6:8])
        try:
            return pd.to_datetime(text).strftime("%Y-%m-%d")
        except Exception:
            return None

    def find_key(cells, keys):
        """判断一行表头 cells 中是否包含 keys 里任一关键字。"""
        for cell in cells:
            for key in keys:
                if key and key in str(cell):
                    return True
        return False

    def pick_column(df, keys):
        """按 keys 顺序在 df.columns 中找第一个包含关键字的列名。"""
        for key in keys:
            for col in df.columns:
                if key in str(col):
                    return col
        return None

    def has_column(df, key):
        return pick_column(df, (key,)) is not None

    def is_credit_transaction(tx_type):
        """中行：根据交易类型判断来账(收入)还是往账(支出)。"""
        text = str(tx_type or "")
        if "来账" in text or "贷" in text:
            return True
        if "往账" in text or "借" in text:
            return False
        return True  # 无法识别时默认按收入处理

    def parse_channel_from_filename(filename):
        """
        从文件名提取收款渠道名。
        去掉日期后缀、账号/活期/明细查询等噪音；蓝乾中信_5701 → 蓝乾中信5701。
        """
        stem = Path(filename).stem
        stem = re.sub(r"\.账号.*$", "", stem)
        stem = re.sub(r"\.活期.*$", "", stem)
        stem = re.sub(r"明细查询.*$", "", stem)
        stem = re.sub(r"[_\.\-]?\d{8}(-\d{8})?$", "", stem)
        stem = stem.replace("_5701", "5701").strip("._- ")
        if "5701" in stem:
            return stem.replace("_", "")
        # 支付宝/微信导出样例：统一渠道名，去掉账号后缀等
        if "支付宝" in stem:
            return "支付宝"
        if "微信" in stem:
            return "微信"
        return stem

    def read_excel_raw(file_path):
        """无表头读 Excel；失败时尝试 read_html（民生 HTML 导出）。"""
        try:
            return pd.read_excel(file_path, header=None)
        except Exception:
            try:
                tables = pd.read_html(str(file_path), encoding="utf-8")
                return tables[0] if tables else None
            except Exception:
                return None

    def detect_header_row(raw, max_rows=35):
        """
        在前 max_rows 行中打分找表头行。
        需日期 + (收支列或对方户名)，或中行「交易类型+交易金额」组合；得分 ≥2 才采纳。
        """
        best_row = None
        best_score = 0
        for i in range(min(max_rows, len(raw))):
            cells = [normalize_header(c) for c in raw.iloc[i].tolist()]
            score = 0
            if find_key(cells, date_keys):
                score += 1
            if find_key(cells, income_keys) or find_key(cells, expense_keys):
                score += 1
            if find_key(cells, counterparty_keys):
                score += 1
            if find_key(cells, ("交易类型",)) and find_key(cells, ("交易金额",)):
                score += 2
            # 微信支付账单：交易时间 + 收/支 + 金额(元)
            if find_key(cells, ("交易时间",)) and find_key(cells, ("收/支",)) and find_key(cells, ("金额",)):
                score += 3
            # 华兴 [HISTORYDETAIL]：交易时间 + 转入/转出金额
            if find_key(cells, ("交易时间",)) and (
                find_key(cells, ("转入金额", "转入额"))
                and find_key(cells, ("转出金额", "转出额"))
            ):
                score += 3
            if score > best_score:
                best_score = score
                best_row = i
        return best_row if best_score >= 2 else None

    def make_record(date_val, counterparty,  summary, income_out, expense_out, channel, flow_type):
        """构造单条标准流水记录。"""
        summary_text = clean_text(summary)
        cp_text = clean_text(counterparty)
        remark_text = ""
        interest_keywords = ["利息", "结息", "存息", "批量结息", "存款利息", "批量结息入账"]
        inter_account_keywords = ["往来款", "往来款项"]
        if any(word in summary_text for word in interest_keywords):
            cp_text = "结息"
        if any(word in summary_text for word in inter_account_keywords):
            remark_text = "往来款"
        return {
            "日期": date_val,
            "对方账户": cp_text,
            "收入": income_out,
            "支出": expense_out,
            "收款渠道": channel,
            "收支类型": flow_type,
            "备注": remark_text,
        }

    def parse_boc_rows(df, channel):
        """中行双语流水：单列交易金额，按交易类型拆收入/支出，对方取付款人或收款人。"""
        col_date = pick_column(df, ("交易日期", "Transaction Date"))
        col_amount = pick_column(df, ("交易金额", "Trade Amount"))
        col_type = pick_column(df, ("交易类型", "Transaction Type"))
        col_payer = pick_column(df, ("付款人名称", "Payer's Name"))
        col_payee = pick_column(df, ("收款人名称", "Payee's Name"))
        if not col_date or not col_amount:
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row.get(col_date))
            amount = normalize_amount(row.get(col_amount))
            if not date_val or not amount:
                continue
            summary = row_summary_text(row, df)
            if is_credit_transaction(row.get(col_type)):
                rows.append(make_record(date_val, row_cell(row, col_payer), summary, amount, None, channel, "收入"))
            else:
                rows.append(make_record(date_val, row_cell(row, col_payee), summary, None, amount, channel, "支出"))
        return rows

    def is_alipay_statement_csv(file_path):
        """识别支付宝官方账务明细 CSV（文件头含 #支付宝账务明细查询）。"""
        try:
            with open(file_path, "rb") as f:
                head = f.read(300).decode("gbk", errors="ignore")
            return "支付宝账务明细" in head or "账务明细列表" in head
        except Exception:
            return False

    def read_alipay_csv(file_path):
        """读取支付宝 CSV：跳过前 4 行说明，兼容 gbk/gb18030。"""
        skiprows = 4
        for encoding in ("utf-8-sig", "gbk", "gb18030"):
            try:
                return pd.read_csv(file_path, encoding=encoding, skiprows=skiprows)
            except Exception:
                continue
        return None

    def read_csv_raw(file_path, skiprows=0):
        """
        读取通用银行 CSV：兼容 utf-16（中行导出）、gbk、utf-8 及 tab/逗号分隔。
        """
        for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "gbk", "gb18030"):
            for sep in ("\t", ",", None):
                kwargs = {"encoding": encoding, "skiprows": skiprows}
                if sep is None:
                    kwargs["sep"] = None
                    kwargs["engine"] = "python"
                else:
                    kwargs["sep"] = sep
                try:
                    df = pd.read_csv(file_path, **kwargs)
                    if df is not None and not df.empty and len(df.columns) >= 3:
                        return df
                except Exception:
                    continue
        return None

    def parse_alipay_csv(file_path, channel):
        """支付宝账务明细 CSV：收入/支出分列，支出列常为负数；遇 # 汇总行停止。"""
        df = read_alipay_csv(file_path)
        if df is None or df.empty:
            return []
        df.columns = [normalize_header(c) for c in df.columns]
        col_date = pick_column(df, ("发生时间", "交易时间", "记账日期"))
        col_income = pick_column(df, ("收入金额（+元）", "收入金额(+元)", "收入"))
        col_expense = pick_column(df, ("支出金额（-元）", "支出金额(-元)", "支出"))
        col_cp = pick_column(df, ("对方账号", "对方账户", "对方户名"))
        if not col_date:
            return []
        rows = []
        for _, row in df.iterrows():
            if str(row.iloc[0]).startswith("#"):
                break
            date_val = normalize_date(row.get(col_date))
            income_val = normalize_amount(row.get(col_income)) if col_income else None
            expense_val = (
                normalize_amount(row.get(col_expense), expense_negative=True)
                if col_expense
                else None
            )
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows


    def parse_csv_file(file_path, channel):
        """CSV 入口：支付宝走专用解析；中行/建行等 CSV 首行即表头，不 skiprows。"""
        if is_alipay_statement_csv(file_path) or "支付宝" in channel:
            return parse_alipay_csv(file_path, channel)
        df = read_csv_raw(file_path, skiprows=0)
        if df is None or df.empty:
            return []
        df.columns = [normalize_header(c) for c in df.columns]
        # 中行双语 CSV（交易类型 + 交易金额）
        if has_column(df, "交易类型") and has_column(df, "交易金额"):
            return parse_boc_rows(df, channel)
        col_date = pick_column(df, date_keys)
        col_income = pick_column(df, income_keys)
        col_expense = pick_column(df, expense_keys)
        col_cp = pick_column(df, counterparty_keys)
        if not col_date or (not col_income and not col_expense):
            return []
        rows = []
        for _, row in df.iterrows():
            if str(row.iloc[0]).startswith("#"):
                break
            date_val = normalize_date(row.get(col_date))
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            income_val = normalize_amount(row_cell(row, col_income)) if col_income else None
            expense_val = normalize_amount(row_cell(row, col_expense)) if col_expense else None
            if income_val is None and expense_val is None:
                continue
            
            if income_val and expense_val:
                if income_val >= expense_val:
                    expense_val = None
                else:
                    income_val = None
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows

    def parse_wechat_rows(df, channel):
        """微信支付账单 xlsx：交易时间、收/支、金额(元)、交易对方。"""
        col_date = pick_column(df, ("交易时间",))
        col_io = pick_column(df, ("收/支",))
        col_amount = pick_column(df, ("金额(元)", "金额"))
        col_cp = pick_column(df, ("交易对方", "对方账户"))
        if not col_date or not col_io or not col_amount:
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row.get(col_date))
            amount = normalize_amount(row.get(col_amount))
            if not date_val or not amount:
                continue
            summary = row_summary_text(row, df)
            io_text = str(row_cell(row, col_io) or "").strip()
            if "收入" in io_text:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, amount, None, channel, "收入"))
            elif "支出" in io_text:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, amount, channel, "支出"))
        return rows

    def parse_huaxing_history_detail(raw, channel):
        """
        华兴网银 [HISTORYDETAIL] 导出：首行标记，第 2 行表头，转入金额=收入、转出金额=支出。
        注意：部分账户流水文件名含「工行」，但文件内容仍是华兴此格式。
        """
        if raw is None or len(raw) < 3:
            return []
        if str(raw.iloc[0, 0]).strip() != "[HISTORYDETAIL]":
            return []
        df = raw.iloc[2:].copy()
        df.columns = [normalize_header(c) for c in raw.iloc[1].tolist()]
        df = df.dropna(how="all")
        if df.empty:
            return []
        col_date = pick_column(df, ("交易时间",))
        col_income = pick_column(df, ("转入金额", "转入额"))
        col_expense = pick_column(df, ("转出金额", "转出额"))
        col_cp = pick_column(df, counterparty_keys)
        if not col_date or (not col_income and not col_expense):
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row_cell(row, col_date))
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            income_val = normalize_amount(row_cell(row, col_income)) if col_income else None
            expense_val = normalize_amount(row_cell(row, col_expense)) if col_expense else None
            if income_val is None and expense_val is None:
                continue
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows

    def parse_bank_file(file_path, channel):
        """单文件入口：csv 走专用解析，xls/xlsx 自动探测表头后分支解析。"""
        if file_path.suffix.lower() == ".csv":
            return parse_csv_file(file_path, channel)
        raw = read_excel_raw(file_path)
        if raw is None or raw.empty:
            return []
        # 华兴 [HISTORYDETAIL]（公户工行2347 等文件名易误解，实为华兴导出）
        if len(raw) >= 1 and str(raw.iloc[0, 0]).strip() == "[HISTORYDETAIL]":
            return parse_huaxing_history_detail(raw, channel)
        header_row = detect_header_row(raw)
        if header_row is None:
            return []
        df = raw.iloc[header_row + 1 :].copy()
        df.columns = [normalize_header(c) for c in raw.iloc[header_row].tolist()]
        df = df.dropna(how="all")
        if df.empty:
            return []
        # 微信支付账单
        if has_column(df, "收/支") and pick_column(df, ("金额(元)", "金额")):
            return parse_wechat_rows(df, channel)
        # 中行格式
        if has_column(df, "交易类型") and has_column(df, "交易金额"):
            return parse_boc_rows(df, channel)
        # 通用双列收支（建行、农行、兴业等）
        col_date = pick_column(df, date_keys)
        col_income = pick_column(df, income_keys)
        col_expense = pick_column(df, expense_keys)
        col_cp = pick_column(df, counterparty_keys)
        if not col_date or (not col_income and not col_expense):
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row.get(col_date))
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            income_val = normalize_amount(row_cell(row, col_income)) if col_income else None
            expense_val = normalize_amount(row_cell(row, col_expense)) if col_expense else None
            if income_val is None and expense_val is None:
                continue
            # 少数行同时有借贷金额时，取较大一侧作为唯一方向
            if income_val and expense_val:
                if income_val >= expense_val:
                    expense_val = None
                else:
                    income_val = None
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows



    # ---------- 扫描文件夹并汇总 ----------
    folder = Path(flow_folder)
    if not folder.is_dir():
        raise ValueError("流水文件夹不存在: %s" % flow_folder)

    records = []
    for pattern in ("*.xls", "*.xlsx", "*.csv"):
        for file_path in sorted(folder.rglob(pattern)):
            channel = parse_channel_from_filename(file_path.name)
            try:
                records.extend(parse_bank_file(file_path, channel))
            except Exception as e:
                print("流水解析失败 [%s]: %s" % (file_path.name, e))
                continue  # 单文件解析失败不影响其余文件

    df_all = pd.DataFrame(records)
    if df_all.empty:
        return df_all, []

    for col in ("日期", "对方账户", "收入", "支出", "收款渠道", "收支类型", "备注"):
        if col not in df_all.columns:
            df_all[col] = None

    df_all = df_all.fillna("")
    for col in ("收入", "支出"):
        df_all[col] = df_all[col].apply(lambda x: "" if x == 0 else x)
    df_all["收款渠道"] = df_all["收款渠道"].astype(str)

    detail_cols = ["日期", "对方账户", "收入", "支出", "收款渠道", "备注"]
    group_cols = ["收款渠道", "收支类型", "日期"]
    result = []

    # ① 先按收款渠道 + 收支类型 + 日期输出支出明细
    df_expense = df_all[df_all["收支类型"] == "支出"].copy()
    if not df_expense.empty:
        for group_keys, group_df in df_expense.groupby(group_cols, sort=False):
            if not isinstance(group_keys, tuple):
                group_keys = (group_keys,)
            record = dict(zip(group_cols, group_keys))
            record["明细"] = group_df[detail_cols].fillna("").to_dict(orient="records")
            result.append(record)

    # ② 再按收款渠道 + 收支类型 + 日期输出收入明细
    df_income = df_all[df_all["收支类型"] == "收入"].copy()
    if not df_income.empty:
        for group_keys, group_df in df_income.groupby(group_cols, sort=False):
            if not isinstance(group_keys, tuple):
                group_keys = (group_keys,)
            record = dict(zip(group_cols, group_keys))
            record["明细"] = group_df[detail_cols].fillna("").to_dict(orient="records")
            result.append(record)

    return df_all, result


def _find_flow_sheet(sheet_names):
    for name in FLOW_SHEET_NAMES:
        if name in sheet_names:
            return name
    return None


def _cell_export(value):
    """写入 Excel 单元格：空值写 None，金额保留数字。"""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        return text
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def append_flows_to_ledger(ledger_path, flow_folder, clear_sheet=False):
    """
    读取流水文件夹（含银行/支付宝/微信），批量写入对账台账「收入流水汇总」sheet。

    对账台账模板列: 日期, 对方账户, 收入, 支出, 收款渠道, 备注,
                    账户类别, 金额, 税额, 是否录入, 结算对象, 核验结果
    后几列留空，由 银行数据处理.py 等后续步骤填充。

    参数:
        ledger_path: 对账台账 xlsx 路径（可为「对账台账模板.xlsx」副本）
        flow_folder: 流水根目录
        clear_sheet: True 时先清空流水 sheet 已有数据行（保留表头）

    返回:
        (本次写入条数, sheet 写入后数据总行数)
    """
    ledger_path = Path(ledger_path)
    if not ledger_path.is_file():
        raise ValueError("台账文件不存在: %s" % ledger_path)

    df_all, _ = get_bank_flows_from_folder(flow_folder)
    if df_all.empty:
        return 0, 0

    wb = load_workbook(ledger_path)
    sheet_name = _find_flow_sheet(wb.sheetnames)
    if not sheet_name:
        raise ValueError(
            "未找到收入流水工作表，当前包含: %s" % list(wb.sheetnames)
        )

    ws = wb[sheet_name]
    if ws.max_row < 1:
        raise ValueError("工作表 [%s] 缺少表头行" % sheet_name)

    if clear_sheet and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    written = 0
    for _, row in df_all.iterrows():
        ws.append([
            _cell_export(row.get("日期")),
            _cell_export(row.get("对方账户")),
            _cell_export(row.get("收入")),
            _cell_export(row.get("支出")),
            _cell_export(row.get("收款渠道")),
            _cell_export(row.get("备注")),
            None,  # 账户类别
            None,  # 金额
            None,  # 税额
            None,  # 是否录入
            None,  # 结算对象
            None,  # 核验结果
        ])
        written += 1

    wb.save(ledger_path)
    total = max(ws.max_row - 1, 0)
    return written, total


if __name__ == "__main__":
    # 本地调试：依次扫描「流水」与「银行流水模板」目录
    base = Path(__file__).resolve().parent.parent
    for folder in (base / "流水", base / "模板文件" / "银行流水模板"):
        if not folder.is_dir():
            continue
        df_all, result = get_bank_flows_from_folder(folder)
        print("=== %s ===" % folder.name)
        print("总流水 %d 条, 分组 %d 组" % (len(df_all), len(result)))
        if not df_all.empty:
            print(df_all[["日期", "对方账户", "收入", "支出", "收款渠道", "收支类型"]].head(5).to_string(index=False))
        for group in result[:5]:
            print(
                "  [%s][%s][%s] %d 条"
                % (group["收款渠道"], group["收支类型"], group["日期"], len(group["明细"]))
            )

    tpl = base / "流程一对账&录入" / "对账台账模板.xlsx"
    flow_dir = base / "流水"
    if tpl.is_file() and flow_dir.is_dir():
        test_ledger = base / "流程一对账&录入" / "_测试写入对账台账.xlsx"
        shutil.copy(tpl, test_ledger)
        w, t = append_flows_to_ledger(test_ledger, flow_dir, clear_sheet=True)
        print("=== 批量写入对账台账测试 ===")
        print("写入 %d 条, 共 %d 条 -> %s" % (w, t, test_ledger.name))
