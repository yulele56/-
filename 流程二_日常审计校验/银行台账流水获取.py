# -*- coding: utf-8 -*-
"""
从流水文件夹读取银行流水明细，区分收入/支出，按收款渠道分组。

支持格式:
    - 建行/农行/兴业/光大/华兴/邮政等 Excel 双列收支
    - 中行双语（交易类型 + 交易金额）
    - 中行 CSV 导出（utf-16 tab 分隔，交易日期 + 收入/支出金额）
    - 华兴网银 [HISTORYDETAIL] 导出（转入金额=收入、转出金额=支出；文件名可能含「工行」）
    - 广州银行等（借方金额 / 贷方金额）
    - 民生等 HTML 伪装 xls（read_html 兜底）

RPA 调用:
    df_all, result = get_bank_flows_from_folder(流水文件夹路径)
    result 顺序: 先各渠道收入明细，再各渠道支出明细

    # 批量写入台账（替代 RPA 逐行填 Excel）:
    written, total = append_flows_to_ledger(台账路径, 流水文件夹路径)
"""

import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

__all__ = ["get_bank_flows_from_folder", "append_flows_to_ledger"]

FLOW_SHEET_NAMES = ("银行收入流水汇总", "银行收入流水", "收入流水汇总")


def get_bank_flows_from_folder(flow_folder):
    """
    读取流水文件夹下银行流水（xls / xlsx / csv），解析并分组。

    返回:
        df_all: 全部流水（空值已 fillna("")）
        result: [{"收款渠道", "收支类型", "日期", "明细"}, ...]
    """
    date_keys = ("记账日期", "交易日期", "交易时间", "发生时间", "日期")
    income_keys = (
        "收入金额（+元）", "收入金额(+元)", "贷方发生额（收入）",
        "贷方发生额/元(收入)", "贷方金额(收入)", "贷方发生额", "贷方金额",
        "存入金额", "收入金额", "收入", "转入金额", "转入额",
    )
    expense_keys = (
        "支出金额（-元）", "支出金额(-元)", "借方发生额（支取）",
        "借方发生额/元(支取)", "借方金额(支出)", "借方发生额", "借方金额",
        "支出金额", "支出", "转出金额", "转出额",
    )
    counterparty_keys = (
        "对方户名", "对方名称", "对方账户", "对方账户名称",
        "对方单位名称", "对方单位", "收款人名称", "付款人名称",
    )
    zhaiyao_keys = (
        "业务摘要", "摘要", "业务类型", "客户附言", "附言", "用途",
    )

    def normalize_header(value):
        if pd.isna(value):
            return ""
        return str(value).strip().replace("\n", "")

    def clean_text(value):
        if isinstance(value, pd.Series):
            value = value.iloc[0] if not value.empty else ""
        if pd.isna(value):
            return ""
        return str(value).strip()

    def row_cell(row, col, default=""):
        if not col:
            return default
        val = row.get(col, default)
        if isinstance(val, pd.Series):
            val = val.iloc[0] if not val.empty else default
        return val

    def row_summary_text(row, df):
        texts = []
        used_cols = set()
        for key in zhaiyao_keys:
            for col in df.columns:
                if col in used_cols:
                    continue
                if key in str(col):
                    used_cols.add(col)
                    text = clean_text(row_cell(row, col))
                    if text:
                        texts.append(text)
                    break
        return " ".join(texts)

    def normalize_amount(value, expense_negative=False):
        if isinstance(value, pd.Series):
            value = value.iloc[0] if not value.empty else None
        if pd.isna(value):
            return None
        text = str(value).strip().replace("\t", "")
        if text in ("", "-", "__", "nan", "NaN", "None"):
            return None
        text = text.replace(",", "").replace(" ", "").replace("￥", "").replace("¥", "").replace("元", "")
        try:
            amount = round(float(text), 2)
        except ValueError:
            return None
        if expense_negative and amount < 0:
            amount = abs(amount)
        return amount if amount > 0 else None

    def normalize_date(value):
        if pd.isna(value):
            return None
        text = str(value).strip()
        if text in ("", "-", "__", "nan"):
            return None
        if re.fullmatch(r"\d{8}\.0+", text):
            text = text.split(".")[0]
        if re.fullmatch(r"\d{8}", text):
            return "%s-%s-%s" % (text[0:4], text[4:6], text[6:8])
        if re.fullmatch(r"\d{8}\s+\d{6}", text):
            return "%s-%s-%s" % (text[0:4], text[4:6], text[6:8])
        try:
            return pd.to_datetime(text).strftime("%Y-%m-%d")
        except Exception:
            return None

    def find_key(cells, keys):
        for cell in cells:
            for key in keys:
                if key and key in str(cell):
                    return True
        return False

    def pick_column(df, keys):
        for key in keys:
            for col in df.columns:
                if key in str(col):
                    return col
        return None

    def has_column(df, key):
        return pick_column(df, (key,)) is not None

    def parse_channel_from_filename(filename):
        stem = Path(filename).stem
        stem = re.sub(r"\.账号.*$", "", stem)
        stem = re.sub(r"\.活期.*$", "", stem)
        stem = re.sub(r"明细查询.*$", "", stem)
        stem = re.sub(r"[_\.\-]?\d{8}(-\d{8})?$", "", stem)
        stem = stem.replace("_5701", "5701").strip("._- ")
        if "5701" in stem:
            return stem.replace("_", "")
        return stem

    def read_excel_raw(file_path):
        try:
            return pd.read_excel(file_path, header=None)
        except Exception:
            try:
                tables = pd.read_html(str(file_path), encoding="utf-8")
                return tables[0] if tables else None
            except Exception:
                return None

    def detect_header_row(raw, max_rows=35):
        best_row, best_score = None, 0
        for i in range(min(max_rows, len(raw))):
            cells = [normalize_header(c) for c in raw.iloc[i].tolist()]
            score = 0
            if find_key(cells, date_keys):
                score += 1
            if find_key(cells, income_keys) or find_key(cells, expense_keys):
                score += 1
            if find_key(cells, counterparty_keys):
                score += 1
            if find_key(cells, ("交易类型",)) and find_key(cells, ("交易金额",)):
                score += 2
            if find_key(cells, ("交易时间",)) and (
                find_key(cells, ("转入金额", "转入额"))
                and find_key(cells, ("转出金额", "转出额"))
            ):
                score += 3
            if score > best_score:
                best_score, best_row = score, i
        return best_row if best_score >= 2 else None

    def make_record(date_val, counterparty, summary, income_out, expense_out, channel, flow_type):
        summary_text = clean_text(summary)
        cp_text = clean_text(counterparty)
        remark_text = ""
        interest_keywords = ["利息", "结息", "存息", "批量结息", "存款利息", "批量结息入账"]
        inter_account_keywords = ["往来款", "往来款项"]
        if any(word in summary_text for word in interest_keywords):
            cp_text = "结息"
        if any(word in summary_text for word in inter_account_keywords):
            remark_text = "往来款"
        return {
            "日期": date_val,
            "对方账户": cp_text,
            "收入": income_out,
            "支出": expense_out,
            "收款渠道": channel,
            "收支类型": flow_type,
            "备注": remark_text,
        }

    def is_credit_transaction(tx_type):
        text = str(tx_type or "")
        if "来账" in text or "贷" in text:
            return True
        if "往账" in text or "借" in text:
            return False
        return True

    def parse_boc_rows(df, channel):
        col_date = pick_column(df, ("交易日期", "Transaction Date"))
        col_amount = pick_column(df, ("交易金额", "Trade Amount"))
        col_type = pick_column(df, ("交易类型", "Transaction Type"))
        col_payer = pick_column(df, ("付款人名称", "Payer's Name"))
        col_payee = pick_column(df, ("收款人名称", "Payee's Name"))
        if not col_date or not col_amount:
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row.get(col_date))
            amount = normalize_amount(row.get(col_amount))
            if not date_val or not amount:
                continue
            summary = row_summary_text(row, df)
            if is_credit_transaction(row.get(col_type)):
                rows.append(make_record(date_val, row_cell(row, col_payer), summary, amount, None, channel, "收入"))
            else:
                rows.append(make_record(date_val, row_cell(row, col_payee), summary, None, amount, channel, "支出"))
        return rows

    def parse_generic_rows(df, channel, expense_negative=False):
        col_date = pick_column(df, date_keys)
        col_income = pick_column(df, income_keys)
        col_expense = pick_column(df, expense_keys)
        col_cp = pick_column(df, counterparty_keys)
        if not col_date or (not col_income and not col_expense):
            return []
        rows = []
        for _, row in df.iterrows():
            if str(row.iloc[0]).startswith("#"):
                break
            date_val = normalize_date(row_cell(row, col_date))
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            income_val = normalize_amount(row_cell(row, col_income)) if col_income else None
            expense_val = (
                normalize_amount(row_cell(row, col_expense), expense_negative=expense_negative)
                if col_expense
                else None
            )
            if income_val is None and expense_val is None:
                continue
            if income_val and expense_val:
                if income_val >= expense_val:
                    expense_val = None
                else:
                    income_val = None
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows

    def read_csv_raw(file_path, skiprows=0):
        for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "gbk", "gb18030"):
            for sep in ("\t", ",", None):
                kwargs = {"encoding": encoding, "skiprows": skiprows}
                if sep is None:
                    kwargs["sep"] = None
                    kwargs["engine"] = "python"
                else:
                    kwargs["sep"] = sep
                try:
                    df = pd.read_csv(file_path, **kwargs)
                    if df is not None and not df.empty and len(df.columns) >= 3:
                        return df
                except Exception:
                    continue
        return None

    def parse_csv_file(file_path, channel):
        df = read_csv_raw(file_path, skiprows=0)
        if df is None or df.empty:
            return []
        df.columns = [normalize_header(c) for c in df.columns]
        if has_column(df, "交易类型") and has_column(df, "交易金额"):
            return parse_boc_rows(df, channel)
        return parse_generic_rows(df, channel, expense_negative=True)

    def parse_huaxing_history_detail(raw, channel):
        if raw is None or len(raw) < 3:
            return []
        if str(raw.iloc[0, 0]).strip() != "[HISTORYDETAIL]":
            return []
        df = raw.iloc[2:].copy()
        df.columns = [normalize_header(c) for c in raw.iloc[1].tolist()]
        df = df.dropna(how="all")
        if df.empty:
            return []
        col_date = pick_column(df, ("交易时间",))
        col_income = pick_column(df, ("转入金额", "转入额"))
        col_expense = pick_column(df, ("转出金额", "转出额"))
        col_cp = pick_column(df, counterparty_keys)
        if not col_date or (not col_income and not col_expense):
            return []
        rows = []
        for _, row in df.iterrows():
            date_val = normalize_date(row_cell(row, col_date))
            if not date_val:
                continue
            summary = row_summary_text(row, df)
            income_val = normalize_amount(row_cell(row, col_income)) if col_income else None
            expense_val = normalize_amount(row_cell(row, col_expense)) if col_expense else None
            if income_val is None and expense_val is None:
                continue
            if income_val and income_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, income_val, None, channel, "收入"))
            elif expense_val and expense_val > 0:
                rows.append(make_record(date_val, row_cell(row, col_cp), summary, None, expense_val, channel, "支出"))
        return rows

    def parse_bank_file(file_path, channel):
        if file_path.suffix.lower() == ".csv":
            return parse_csv_file(file_path, channel)
        raw = read_excel_raw(file_path)
        if raw is None or raw.empty:
            return []
        if len(raw) >= 1 and str(raw.iloc[0, 0]).strip() == "[HISTORYDETAIL]":
            return parse_huaxing_history_detail(raw, channel)
        header_row = detect_header_row(raw)
        if header_row is None:
            return []
        df = raw.iloc[header_row + 1 :].copy()
        df.columns = [normalize_header(c) for c in raw.iloc[header_row].tolist()]
        df = df.dropna(how="all")
        if df.empty:
            return []
        if has_column(df, "交易类型") and has_column(df, "交易金额"):
            return parse_boc_rows(df, channel)
        return parse_generic_rows(df, channel)

    # ---------- 扫描文件夹（银行 xls / xlsx / csv） ----------
    folder = Path(flow_folder)
    if not folder.is_dir():
        raise ValueError("流水文件夹不存在: %s" % flow_folder)

    records = []
    for pattern in ("*.xls", "*.xlsx", "*.csv"):
        for file_path in sorted(folder.rglob(pattern)):
            if file_path.name.startswith("~$"):
                continue
            channel = parse_channel_from_filename(file_path.name)
            try:
                records.extend(parse_bank_file(file_path, channel))
            except Exception as e:
                print("流水解析失败 [%s]: %s" % (file_path.name, e))
                continue

    df_all = pd.DataFrame(records)
    if df_all.empty:
        return df_all, []

    for col in ("日期", "对方账户", "收入", "支出", "收款渠道", "收支类型", "备注"):
        if col not in df_all.columns:
            df_all[col] = None

    df_all = df_all.fillna("")
    for col in ("收入", "支出"):
        df_all[col] = df_all[col].apply(lambda x: "" if x == 0 else x)
    df_all["收款渠道"] = df_all["收款渠道"].astype(str)

    detail_cols = ["日期", "对方账户", "收入", "支出", "收款渠道", "备注"]
    group_cols = ["收款渠道", "收支类型", "日期"]
    result = []

    for flow_type in ("收入", "支出"):
        part = df_all[df_all["收支类型"] == flow_type]
        if part.empty:
            continue
        for group_keys, group_df in part.groupby(group_cols, sort=False):
            if not isinstance(group_keys, tuple):
                group_keys = (group_keys,)
            record = dict(zip(group_cols, group_keys))
            record["明细"] = group_df[detail_cols].fillna("").to_dict(orient="records")
            result.append(record)

    return df_all, result


def _find_flow_sheet(sheet_names):
    for name in FLOW_SHEET_NAMES:
        if name in sheet_names:
            return name
    return None


def _cell_export(value):
    """写入 Excel 单元格：空值写 None，金额保留数字。"""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        return text
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def append_flows_to_ledger(ledger_path, flow_folder, clear_sheet=False):
    """
    读取流水文件夹，批量追加写入台账「银行收入流水汇总」sheet（openpyxl 只改该 sheet）。

    参数:
        ledger_path: 审计/对账台账 xlsx 路径
        flow_folder: 银行流水根目录
        clear_sheet: True 时先清空流水 sheet 已有数据行（保留第 1 行表头）

    返回:
        (本次写入条数, sheet 写入后数据总行数)
    """
    ledger_path = Path(ledger_path)
    if not ledger_path.is_file():
        raise ValueError("台账文件不存在: %s" % ledger_path)

    df_all, _ = get_bank_flows_from_folder(flow_folder)
    if df_all.empty:
        return 0, 0

    wb = load_workbook(ledger_path)
    sheet_name = _find_flow_sheet(wb.sheetnames)
    if not sheet_name:
        raise ValueError(
            "未找到银行收入流水工作表，当前包含: %s" % list(wb.sheetnames)
        )

    ws = wb[sheet_name]
    if ws.max_row < 1:
        raise ValueError("工作表 [%s] 缺少表头行" % sheet_name)

    if clear_sheet and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    written = 0
    for _, row in df_all.iterrows():
        income = row.get("收入", "")
        expense = row.get("支出", "")
        ws.append([
            _cell_export(row.get("日期")),
            _cell_export(row.get("对方账户")),
            _cell_export(income),
            _cell_export(expense),
            _cell_export(row.get("收款渠道")),
            _cell_export(row.get("备注")),
            None,
            None,
            None,
        ])
        written += 1

    wb.save(ledger_path)
    total = max(ws.max_row - 1, 0)
    return written, total


if __name__ == "__main__":
    base = Path(__file__).resolve().parent.parent
    for folder in (base / "流水", base / "模板文件" / "银行流水模板"):
        if not folder.is_dir():
            continue
        df_all, result = get_bank_flows_from_folder(folder)
        print("=== %s ===" % folder.name)
        print("总流水 %d 条, 分组 %d 组" % (len(df_all), len(result)))
        if not df_all.empty:
            print(df_all[["日期", "对方账户", "收入", "支出", "收款渠道", "收支类型"]].head(5).to_string(index=False))

    tpl = base / "流程二_日常审计校验" / "审计台账模板.xlsx"
    flow_dir = base / "流水"
    if tpl.is_file() and flow_dir.is_dir():
        test_ledger = base / "流程二_日常审计校验" / "_测试写入台账.xlsx"
        shutil.copy(tpl, test_ledger)
        w, t = append_flows_to_ledger(test_ledger, flow_dir, clear_sheet=True)
        print("=== 批量写入测试 ===")
        print("模板: %s" % tpl.name)
        print("写入 %d 条, 台账共 %d 条" % (w, t))
